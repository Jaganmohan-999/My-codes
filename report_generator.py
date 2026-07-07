from openpyxl import Workbook
from openpyxl.styles import Font

# =========================
# 🔧 CONFIG (EDIT HERE ONLY)
# =========================

use_cases = [
    "Person Detection in Unauthorized time",
    "No one at Cash Counter",
    "Employees in Group",
    "Store Opening Time",
    "Store Closing Time",
    "Customer Unattended"
]

locations = [
   "RS KPHB",
   "SI Ameerpet",
   	"SI Patny",
    "SI Kothapet",
    "SI Madinaguda",
    "RS Dilsukhnagar",
    "RS Chandanagar",
    "SI Hanamkonda",
    "SI Vizianagaram",
    "SI Suchitra",
    "SI Kukatpally",
    "RS Gachibowli",
    "SI Uppal",
    "SI Rajahmundry",
    "RS Ameerpet"
]

dates = [
    "Saturday",
    "Sunday"
]

output_file = "usecase_precision_report.xlsx"

# =========================
# 📊 CREATE WORKBOOK
# =========================
wb = Workbook()

# =========================
# 🧾 DATA ENTRY SHEET
# =========================
ws = wb.active
ws.title = "Data Entry"

headers = [
    "Date", "Location", "Use Case",
    "Total Reviewed", "TP", "FP", "Precision"
]

ws.append(headers)

# Bold header
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = Font(bold=True)

row = 2

# Generate rows
for date in dates:
    for loc in locations:
        for uc in use_cases:
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=loc)
            ws.cell(row=row, column=3, value=uc)

            # Precision formula
            ws.cell(row=row, column=7, value=f"=IF((E{row}+F{row})=0,0,E{row}/(E{row}+F{row}))")

            row += 1

# =========================
# 📊 LOCATION SUMMARY
# =========================
ws2 = wb.create_sheet("Location Summary")

ws2.append(["Location", "Total TP", "Total FP", "Precision"])

for col in range(1, 5):
    ws2.cell(row=1, column=col).font = Font(bold=True)

for i, loc in enumerate(locations, start=2):
    ws2.cell(row=i, column=1, value=loc)
    ws2.cell(row=i, column=2, value=f'=SUMIF(\'Data Entry\'!B:B,"{loc}",\'Data Entry\'!E:E)')
    ws2.cell(row=i, column=3, value=f'=SUMIF(\'Data Entry\'!B:B,"{loc}",\'Data Entry\'!F:F)')
    ws2.cell(row=i, column=4, value=f"=IF((B{i}+C{i})=0,0,B{i}/(B{i}+C{i}))")

# =========================
# 📊 USECASE OVERALL SUMMARY
# =========================
ws3 = wb.create_sheet("Usecase Overall")

ws3.append(["Use Case", "Total TP", "Total FP", "Overall Precision"])

for col in range(1, 5):
    ws3.cell(row=1, column=col).font = Font(bold=True)

for i, uc in enumerate(use_cases, start=2):
    ws3.cell(row=i, column=1, value=uc)
    ws3.cell(row=i, column=2, value=f'=SUMIF(\'Data Entry\'!C:C,"{uc}",\'Data Entry\'!E:E)')
    ws3.cell(row=i, column=3, value=f'=SUMIF(\'Data Entry\'!C:C,"{uc}",\'Data Entry\'!F:F)')
    ws3.cell(row=i, column=4, value=f"=IF((B{i}+C{i})=0,0,B{i}/(B{i}+C{i}))")

# =========================
# 💾 SAVE FILE
# =========================
wb.save(r"D:\Usecase_report\usecase_precision_report.xlsx")

print(f"✅ Report generated: {output_file}")